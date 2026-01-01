import re
import os
import csv
from openpyxl import load_workbook
from playwright.sync_api import Page, expect
from config import USERNAME, PASSWORD, URL

def test_download_citation_for_company(page: Page, tmp_path):

    # Login
    page.goto(URL, wait_until="domcontentloaded", timeout=60000)
    page.get_by_role("textbox", name="Username").fill(USERNAME)
    page.get_by_role("textbox", name="Password").fill(PASSWORD)
    page.get_by_role("button", name="Sign In").click()

    # Navigate to Compliance
    page.get_by_role("link", name="Compliance").click()
    expect(page.get_by_text("Compliance").nth(1)).to_be_visible(timeout=20000)

    # Select Company
    page.get_by_role("combobox").nth(0).select_option("115")

    # Select Mine
    page.get_by_role("combobox").nth(1).select_option("95")

    expect(page.locator("table")).to_be_visible(timeout=30000)

    # ---- Download ----
    with page.expect_download() as download_info:
        page.get_by_role("button", name=re.compile("CSV|Download", re.I)).click()
        page.get_by_role("button", name="Company").click()

    download = download_info.value

    # ------------------------------------------
    # Save to project download folder
    # ------------------------------------------
    download_dir = os.path.join(os.getcwd(), "download", "company_citation")
    os.makedirs(download_dir, exist_ok=True)

    file_path = os.path.join(download_dir, download.suggested_filename)
    download.save_as(file_path)

    print(f"Downloaded file saved at: {file_path}")

    assert os.path.exists(file_path)


   # Read downloaded file & print the column names
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    print("XLSX Headers:", headers)
    assert len(headers) > 1, "XLSX file seems invalid!"





def test_download_citation_for_current_mine(page: Page, tmp_path):

    # Login
    page.goto(URL, wait_until="domcontentloaded", timeout=60000)
    page.get_by_role("textbox", name="Username").fill(USERNAME)
    page.get_by_role("textbox", name="Password").fill(PASSWORD)
    page.get_by_role("button", name="Sign In").click()

    # Navigate to Compliance
    page.get_by_role("link", name="Compliance").click()
    expect(page.get_by_text("Compliance").nth(1)).to_be_visible(timeout=20000)

    # Select Company
    page.get_by_role("combobox").nth(0).select_option("115")

    # Select Mine
    page.get_by_role("combobox").nth(1).select_option("95")

    expect(page.locator("table")).to_be_visible(timeout=30000)

    # ---- Download ----
    with page.expect_download() as download_info:
        page.get_by_role("button", name=re.compile("CSV|Download", re.I)).click()
        page.get_by_role("button", name="Current Mine").click()

    download = download_info.value

    # ------------------------------------------
    # Save to project download folder
    # ------------------------------------------
    download_dir = os.path.join(os.getcwd(), "download", "current_mine_citation")
    os.makedirs(download_dir, exist_ok=True)

    file_path_current_mine = os.path.join(download_dir, download.suggested_filename)
    download.save_as(file_path_current_mine)

    print(f"Downloaded file saved at: {file_path_current_mine}")

    assert os.path.exists(file_path_current_mine)


   # Read downloaded file & print the column names
    workbook = load_workbook(filename=file_path_current_mine)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    print("XLSX Headers:", headers)
    assert len(headers) > 1, "XLSX file seems invalid!"
